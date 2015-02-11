from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import re

#global variables
F1 = 'C:/Users/cwj/Desktop/A_ CC for TS.xlsx'
F2 = 'E:/Excel/genotyping{}.xlsx'.format(datetime.datetime.now().year)
MLG = []#mixed litter groups for setting mixed_index
Primer_dict = {'Cre': {'Tg': 'Cre'},
'H3f3a': {'-': '3Adel4', '+': '3Afl1', 'aNG-': '3Adel2', 'aNGfl': '3Adel2', 'Avi-LZ': 'LacZ', 'fl': '3Afl2', 'G-': '(VISUAL)', 'Gfl': '(VISUAL)'},
'H3f3b': {'-': '3Bdel4', '+': '3Bfl1', 'aNt-': '3Bdel3n2', 'aNtfl': '3Bdel5n', 'Avi-LZ': 'LacZ', 'fl': '3Bfl2', 'tdT-': '(VISUAL)', 'tdTfl': '(VISUAL)'},
'Pgk-FLPo': {'Tg': 'FLPo'},
'R26': {'+': 'R26', 'CA-BA': 'BirA', 'CA-tdT-DBA': '(VISUAL)', 'PhiC': 'PhiC', 'CA-tdT-BA': '(VISUAL)', 'Flpe': 'FLP', 'LZ': 'LacZ',
'CA-DBA': 'BirA', 'tdT': 'tdT', 'tdT-BA': '(VISUAL)', 'tdT-DBA': '(VISUAL)'},
'Trp53': {'-': 'p53del', '+': 'p53WT'}}
GT_priority_dict = {'(undefined)': -1, '(VISUAL)': 0, 'p53WT': 1, 'p53del': 2, 'Cre': 3}
GT_date = datetime.datetime.now()
wb = None
litters = []
scores = {}

#regular expression    
death = re.compile('-\d+[dfm]')
mix = re.compile('[Mm]. w/ L\d+')
fo = re.compile('[Ff]o.* [Bb]y L\d+')

#Litter class
class Litter(object):
    def __init__(self, litterID, DadID, DadGT, MomID, MomGT, DOB, size, clip_date, PS):
        self.litterID = litterID
        self.DadID = DadID
        self.DadGT = DadGT
        self.MomID = MomID
        self.MomGT = MomGT
        self.DOB = DOB
        self.size = size
        self.clip_date = clip_date
        if PS:
            self.PS = PS
        else:
            self.PS = ''
        self.realsize = real_size(self.size, self.PS)
        self.related_L = related(self.PS)
        self.DadGT_dict = GT_parser(self.DadGT)
        self.MomGT_dict = GT_parser(self.MomGT)
        self.GT_alleles = GT_need(self.DadGT_dict, self.MomGT_dict)
        self.first_toe = 1
        self.mixed_index = None
        self.primer_sets = find_primers(self.GT_alleles)
        self.first = None
    
    def __str__(self):
        return str([self.litterID, self.first, self.primer_sets, self.GT_alleles,self.DOB, self.mixed_index])
#        return str([self.litterID, self.DadID, self.DadGT, self.MomID, self.MomGT, self.DOB, self.size, self.PS, self.realsize, self.GT_alleles, self.mixed_index])
    
    def update(self, first_toe = None, sort_dict = None, mixed_index = None, primer_sets = None, first = None, GT_alleles = None, realsize = None):
        if first_toe != None:
            self.first_toe = first_toe
        if sort_dict != None:
            self.GT_alleles.sort(key = sort_dict.get)
        if mixed_index != None:
            self.mixed_index = mixed_index
        if primer_sets != None:
            self.primer_sets = primer_sets
        if first != None:
            self.first = first
        if GT_alleles != None:
            self.GT_alleles = GT_alleles
        if realsize != None:
            self.realsize = realsize
    
    def get_info(self, item):
        if item == 'related_L':
            return self.related_L
        elif item == 'litterID':
            return self.litterID
        elif item == 'DadGT':
            return self.DadGT
        elif item == 'MomGT':
            return self.MomGT
        elif item == 'DadGT_dict':
            return self.DadGT_dict
        elif item == 'MomGT_dict':
            return self.MomGT_dict
        elif item == 'mixed_index':
            return self.mixed_index
        elif item == 'primer_sets':
            return self.primer_sets
        elif item == 'realsize':
            return self.realsize
        elif item == 'first':
            return self.first
        elif item == 'GT_alleles':
            return self.GT_alleles
        elif item == 'DOB':
            return self.DOB
        elif item == 'clip_date':
            return self.clip_date
        else:
            return None
        
    def write_list(self, ws, r):
        ws.cell(row = r,column = 0).value = self.litterID
        ws.cell(row = r,column = 1).value = self.DadID
        ws.cell(row = r,column = 2).value = self.DadGT
        ws.cell(row = r,column = 3).value = self.MomID
        ws.cell(row = r,column = 4).value = self.MomGT
        ws.cell(row = r,column = 5).value = self.DOB
        ws.cell(row = r,column = 6).value = self.size
        ws.cell(row = r,column = 7).value = self.realsize
        ws.cell(row = r,column = 8).value = self.PS
        ws.cell(row = r,column = 9).value = self.mixed_index
        ws.cell(row = r,column = 10).value = str(self.primer_sets)
        ws.cell(row = r,column = 11).value = self.first
        
    def make_TS(self, ws, r, Tube):
        R = r
        toe_cut = self.first_toe
        tube_num = Tube
        ws.cell(row = R,column = 0).value = self.litterID
        ws.cell(row = R,column = 1).value = self.DadID
        ws.cell(row = R,column = 2).value = self.DadGT
        ws.cell(row = R,column = 3).value = self.MomID
        ws.cell(row = R,column = 4).value = self.MomGT
        ws.cell(row = R,column = 5).value = str(self.DOB.month) + '/' + str(self.DOB.day)
        for x in range(R, R + self.realsize):
            ws.cell(row = R,column = 6).value = toe_cut
            ws.cell(row = R,column = 9).value = self.first
            if self.first:
                ws.cell(row = R,column = 10).value = tube_num
            R += 1
            tube_num += 1
            toe_cut = next_toe(toe_cut)
        return R, tube_num
        


# helper functions:
def special_case(litters):# deal w/ FP turned on in F1 alleles
    for lit in litters:
        GT_alleles = set(lit.get_info('GT_alleles'))
        primers = set(lit.get_info('primer_sets'))
        FP_alleles = {('H3f3a', 'aNGfl'), ('H3f3a', 'aNG-'), ('H3f3b', 'aNt-'), ('H3f3b', 'aNtfl')}
        if ('R26', 'PhiC') in GT_alleles or 'PhiC' in lit.get_info('MomGT_dict').get('R26', []) or 'PhiC' in lit.get_info('DadGT_dict').get('R26', []) :
            F = GT_alleles.intersection(FP_alleles)
            if len(F) > 0:
                primers.add('(VISUAL)')
                GT_alleles.difference_update(FP_alleles)
                primers.difference_update(set(find_primers(F)))
        lit.update(GT_alleles = list(GT_alleles))
        lit.update(primer_sets = list(primers))
    
    
def find_primers(GT_alleles):
    primers = set()
    for alle in GT_alleles:
        p_dict = Primer_dict.get(alle[0], None)
        if p_dict:
            primers.add(p_dict.get(alle[1], '(undefined)'))
        else:
            primers.add('(undefined)')
    primers = list(primers)
    return primers

def Top(lit):
    primers = lit.get_info('primer_sets')
    groups = {} # priority groups
    for p in primers:
        num = GT_priority_dict.get(p, 1000)
        groups[num] = groups.get(num, []) + [p]
    if groups:
        return groups[min(groups)]
    else:
        return ''

def sort_key(lit):
    first = lit.get_info('first')
    v = scores.values()
    if first == '(VISUAL)': #no PCR needed, second bottom
        a = min(v) - 1
    elif first == '(undefined)': #Error, put on top
        a = max(v) + 1
    elif first:
        a = scores[lit.get_info('first')] #popularity sort
    else: # No need to genotype, very bottom
        a = min(v) - 2
    return (a, len(lit.get_info('primer_sets')), lit.get_info('DadGT'), lit.get_info('MomGT'), lit.get_info('mixed_index'), datetime.datetime.now() - lit.get_info('DOB'))

def litters_sort(L):
    global scores
    scores = {}  
    for lit in L:
        for p in Top(lit):
            scores[p] = scores.get(p, 0) + lit.get_info('realsize')
    for lit in L:
        top = Top(lit)
        if top:
            top.sort(key = scores.get, reverse = True)
            lit.update(first = top[0])
    L.sort(key = sort_key, reverse = True)

def mixed_index(litters):
    #identify who & who are in the same cage
    litters_copy = list(litters)
    for a in litters:
        if a.get_info('mixed_index') == None:
            a.update(mixed_index = len(MLG))
            MLG.append([a])
        a_index = a.get_info('mixed_index')
        a_group = set(a.get_info('related_L'))
        a_group.add(a.get_info('litterID'))
        litters_copy.remove(a)
        for b in litters_copy:
            if b.get_info('mixed_index') == None:
                b_group = set(b.get_info('related_L'))
                b_group.add(b.get_info('litterID'))
                if len(a_group.intersection(b_group)) > 0:
                    MLG[a_index].append(b)
                    b.update(mixed_index = a_index)
    MLG_copy = list(MLG)
    for group in MLG_copy:
        if len(group) == 1:
            MLG.remove(group)
            group[0].update(mixed_index = -1)
    for group in MLG:
        i = MLG.index(group)
        for lit in group:
            lit.update(mixed_index = i)
 

def next_toe(toe_cut):
    t = toe_cut
    if t == 10:
        return 21
    elif t % 10 != 9:
        return t + 1
    elif t == 9:
        return 10
    elif t == 99:
        return 1
    else:
        return t + 2

def first_toe_cal(N):
    N = N + 1
    if N <= 10:
        return N
    if N == 82:
        return 99
    elif N > 82:
        return first_toe_cal(N % 82 - 1)
    else:
        M = (N - 10) // 9 + 2
        L = (N - 10) % 9
        return 10 * M + L

def real_size(size, PS):
    D = death.findall(PS)
    real = size
    for d in D:
        real += int(d[:-1])
    return real
    
def related(PS):
    m = mix.findall(PS)
    n = fo.findall(PS)
    all = set()
    if m:
        all.add(int(m[-1].split('L')[-1]))
    if n:
        all.add(int(n[-1].split('L')[-1]))
    return all
        
def GT_parser(GT):
    g = GT.split(';') # list of [genes_allele/allele,...]
    g1 = {}
    for i in g:
        g2 = i.split('_') # list of [gene, allele/allele]
        if len(g2) > 1:
            g1[g2[0].strip()] = g2[1].split('/')
    return g1 # dict of {gene:[allele/allele],...}

def GT_need(DadGT_dict, MomGT_dict):
    D = DadGT_dict
    M = MomGT_dict
    n = dict()  #needed to do genotyping
    for i in D:
        if D[i][0] != D[i][1]: #if homozygous, no need to genotype
            n[i] = D[i]
    for i in M:
        if i in n:#if gene has been added.
            if M[i][0] == M[i][1]: #if Mom is homo, dad isn't
                if M[i][0] in n[i]:
                    n[i].remove(M[i][0]) #Remove the allele that doesn't need to be genotyped
            else:
                for a in M[i]:
                    if a not in n[i]:
                        n[i].append(a) #Add alleles not listed
        else:
            if M[i][0] != M[i][1]:
                if i in D:#Dad is homo for this gene
                    for a in M[i]:
                        if a != D[i][0]:
                            n[i] = n.get(i,[]) + [a] #add the different allele(s)
                else:
                    n[i] = M[i]
    for i in n:
        if len(n[i]) > 1:
            for a in n[i]:
                if a == '+' or a == '0':
                    n[i].remove(a)
    N = set()
    for i in n:
        for a in n[i]:
            N.add((i, a))
    N_copy = set(N)
    for i in N_copy:
        if 'Cre' in i[0] or 'GCer' in i[0]: #deal w/ Cre transgenes
            if i[1] == 'Tg':
                N.discard(i)
                N.add(('Cre','Tg'))
            elif i[1] == '0':
                N.discard(i)
    N = list(N)
    return N

def set_GT_date():
    q = input('Are you doing genotyping today? (y/n)')
    if q in ['Y', 'y', '']:
        return datetime.datetime.now()
    else:
        q = input('Tomorrow?')
        if q in ['Y', 'y', '']:
            return datetime.datetime.now() + datetime.timedelta(1)
        else:
            try:
                q1 = int(input('How many days later?'))
                if q1 < 0:
                    raise ValueError
                return datetime.datetime.now() + datetime.timedelta(q1)
            except ValueError:
                print('Wrong value!')
                return set_GT_date()

def collect_litters():
    global litters   
    GT = wb.get_sheet_by_name('Genotype')
    for r in range(2, len(GT.row_dimensions), 4):
        #lit = [0 litter ID, 1 Dad, 2 Dad GT, 3 Mom, 4 Mom GT, 5 DOB,6 Litter size, 7 Clip toes, 8 PS]
        lit = [GT.cell(row = r, column = 0).value]#0 Litter ID
        lit.append(GT.cell(row = r, column = 1).value)#1 Dad
        lit.append(GT.cell(row = r + 3, column = 7).value)#2 Dad GT
        lit.append(GT.cell(row = r, column = 2).value)#3 Mom
        lit.append(GT.cell(row = r + 2, column = 7).value)#4 Mom GT
        lit.append(GT.cell(row = r, column = 3).value)#5 DOB
        lit.append(GT.cell(row = r, column = 4).value)#6 Litter size
        lit.append(GT.cell(row = r, column = 5).value)#7 Clip toes
        lit.append(GT.cell(row = r, column = 6).value)#8 PS
        litters.append(Litter(lit[0], lit[1], lit[2], lit[3], lit[4], lit[5], lit[6], lit[7], lit[8]))

def filter_litters(litters):
    global GT_date
    GT_date = set_GT_date()
    litters_copy = litters[:]
    for lit in litters_copy:
        if lit.get_info('clip_date') != None or GT_date - lit.get_info('DOB') < datetime.timedelta(6) or lit.get_info('realsize') <= 0:
            litters.remove(lit)


def job():
    global F1
    print('"' + F1 + '" is the default file for output and reading raw data input.')
    print('Enter "s" to change the file and path.')
    print('Enter "1" to generate litter list for review.')
    print('Enter "2" to generate toe sheet from the reviewed list.')
    print('Enter "3" to generate genotyping plate.')
    print('Enter "4" to calculate reagent amounts from edited "plate" sheet.')
    print('Enter "5" to match qPCR data (in "data" sheet) with plate (in "plate" sheet).')
    print('Enter "6" to match reviewed qPCR results (in "plate_match" sheet) with toe sheet.')
    print('Enter "c" to cancel.')
    ans = input('What do you want to do?\n>')
    print()
    print()
    if ans == 's' or ans == 'S':
        F1 = input('Enter the full path and file name for the data input file:')
        print()
        return job()
    if ans == 'c' or ans == 'C':
        return ans
    if ans not in '123456' or ans == '':
        print('Wrong value!\n')
        return job()
    ans = int(ans)
    if ans < 1 or ans > 6:
        print('Wrong value!\n')
        return job()
    return ans

def make_list(ws, litters = litters):
    ws.cell(row = 0, column = 0).value = 'litterID'
    ws.cell(row = 0, column = 1).value = 'DadID'
    ws.cell(row = 0, column = 2).value = 'DadGT'
    ws.cell(row = 0, column = 3).value = 'MomID'
    ws.cell(row = 0, column = 4).value = 'MomGT'
    ws.cell(row = 0, column = 5).value = 'DOB'
    ws.cell(row = 0, column = 6).value = 'size'
    ws.cell(row = 0, column = 7).value = 'realsize'
    ws.cell(row = 0, column = 8).value = 'PS'
    ws.cell(row = 0, column = 9).value = 'mixed_index'
    ws.cell(row = 0, column = 10).value = 'primer_sets'
    ws.cell(row = 0, column = 11).value = 'first'
    R = 1
    for lit in litters:
        lit.write_list(ws, R)
        R += 1
    ws.cell(row = R, column = 6).value = 'Total'
    total = 0
    for i in range(1, R):
        total += ws.cell(row = i, column = 7).value
    ws.cell(row = R, column = 7).value = total
    
def first_toe_num():
    new_MLG = {}
    for lit in litters:
        if lit.get_info('mixed_index') != -1:
            new_MLG[lit.get_info('mixed_index')] = new_MLG.get(lit.get_info('mixed_index'), []) + [lit]
    for group in new_MLG:
        N = 0
        for lit in new_MLG[group]:
            lit.update(first_toe_cal(N))
            N += lit.get_info('realsize')

def generate_TS():
    TS = wb.create_sheet(index = 0, title = 'TS')
    TS.cell(row = 0, column = 0).value = 'litterID'
    TS.cell(row = 0, column = 1).value = 'DadID'
    TS.cell(row = 0, column = 2).value = 'DadGT'
    TS.cell(row = 0, column = 3).value = 'MomID'
    TS.cell(row = 0, column = 4).value = 'MomGT'
    TS.cell(row = 0, column = 5).value = 'DOB'
    TS.cell(row = 0, column = 6).value = 'Toe#'
    TS.cell(row = 0, column = 7).value = 'C.C.'
    TS.cell(row = 0, column = 8).value = 'Sex'
    TS.cell(row = 0, column = 9).value = 'first'
    TS.cell(row = 0, column = 10).value = 'Tube#'
    TS.cell(row = 0, column = 11).value = 'PS'
    row = 1
    tube = 1
    for lit in litters:
        row, tube = lit.make_TS(ws = TS, r = row, Tube = tube)
    try:
        wb.save(F1)
        print('"TS" has been generated.')
        return True
    except PermissionError:
        print('Please close "' + F1 + '" for file writing!')
        print('*' * 20)
        print()
        return False

def updater(litter_map, GT1, field, data_cols):
    '''updates individual litter's infomration after list was reviewed'''
    if data_cols.get(field, None) != None:
        if field == 'realsize':
            for r in range(1, len(GT1.row_dimensions)):
                lit = litter_map.get(GT1.cell(row = r, column = data_cols['litterID']).value, None)
                if lit != None:
                    val = GT1.cell(row = r, column = data_cols[field]).value
                    lit.update(realsize = val)
        elif field == 'mixed_index':
            for r in range(1, len(GT1.row_dimensions)):
                lit = litter_map.get(GT1.cell(row = r, column = data_cols['litterID']).value, None)
                if lit != None:
                    val = GT1.cell(row = r, column = data_cols[field]).value
                    lit.update(mixed_index = val)
        elif field == 'first':
            for r in range(1, len(GT1.row_dimensions)):
                lit = litter_map.get(GT1.cell(row = r, column = data_cols['litterID']).value, None)
                if lit != None:
                    val = GT1.cell(row = r, column = data_cols[field]).value
                    lit.update(first = val)

def list_updater(GT1):
    '''updates litter list and individual litter info (through calling updater)'''
    global litters
    data_cols = {}
    new_order = []
    litter_map = {}
    for c in range(len(GT1.column_dimensions)):
        data_name = GT1.cell(row = 0, column = c).value
        if data_name == None:
            return False
        data_cols[data_name] = c
    if data_cols.get('litterID', None) != None:
        for lit in litters:
            litter_map[lit.get_info('litterID')] = lit
        for r in range(1, len(GT1.row_dimensions)):
            l = GT1.cell(row = r, column = data_cols['litterID']).value
            if l != None:
                new_order.append(l)
            else:
                if r != len(GT1.row_dimensions) - 1:
                    return False
        litters = []
        for l in new_order:
            litters.append(litter_map[l])
        updater(litter_map, GT1,'realsize', data_cols)
        updater(litter_map, GT1,'first', data_cols)
        updater(litter_map, GT1,'mixed_index', data_cols)
    first_toe_num()
    return True

def reader(TS, Range):
    for row in TS.iter_rows(Range):
        for cell in row:
            yield cell.internal_value

def read_cols(TS, two = True):
    if two == True:
        FTP = [] #First & Tube number pairs
        b = reader(TS, 'J2:L999')
        end = False
        while end == False:
            f = next(b, None)
            t = next(b, None)
            if f and t:
                FTP.append((f, t))
            else:
                end = True
        return FTP
    else:
        Tube = []
        a = reader(TS, 'K2:L999')
        end = False
        while end == False:
            s = next(a, None)
            if s:
                Tube.append(s)
            else:
                end = True
        return Tube
        


def generate_plate(TS):
    FTP = read_cols(TS)
    plate = wb.create_sheet(0, 'plate')
    for col in ['AB','CD','EF','GH','IJ','KL','MN','OP','QR','ST','UV','WX']:
        if len(FTP) == 0:
            break
        for row in '12345678':
            if len(FTP) == 0:
                break
            pair = FTP.pop(0)
            if pair[0] != '(VISUAL)':
                plate.cell(col[0] + row).value = pair[0]
                plate.cell(col[1] + row).value = pair[1]
    p = range(len(FTP) // 96 + 2)
    for i in p:
        if i != 0:
            I = str(i + 1)
        else:
            I = ''
        c = 1
        for col in 'ACEGIKMOQSUW':
            plate.cell(col + I + '9').value = c
            c += 1
    r = 2
    while len(FTP) > 0:
        R = str(r)
        for col in ['AB','CD','EF','GH','IJ','KL','MN','OP','QR','ST','UV','WX']:
            if len(FTP) == 0:
                break
            for row in '12345678':
                if len(FTP) == 0:
                    break
                pair = FTP.pop(0)
                if pair[0] != '(VISUAL)':
                    plate.cell(col[0] + R + row).value = pair[0]
                    plate.cell(col[1] + R + row).value = pair[1]
        r += 1
        
    try:
        wb.save(F1)
        print('Plate has been generated!')
        print()
        return True
    except PermissionError:
        print('Please close "' + F1 + '" for file writing!')
        print('*' * 20)
        print()
        return False

def well_ID():
    for j in [1,2,3,4,5,6,7,8,9,10,11,12]:
        for i in 'ABCDEFGH':
            yield i + str(j).rjust(2, '0')
            


def read_plate(plate, extra = False):
    if extra == False:
        qFTP = {} #qPCR result and First-Tube number pair dict
        id = well_ID()
        for col in ['AB','CD','EF','GH','IJ','KL','MN','OP','QR','ST','UV','WX']:
            for row in '12345678':
                qFTP[next(id)] = [plate.cell(col[0] + row).value, plate.cell(col[1] + row).value]
        return qFTP
    else: #look into 2nd plate and more
        extra = {}
        N = 2
        while plate.cell('A' + str(N) + '1').value != None:
            n = str(N)
            for col in 'ACEGIKMOQSUW':
                for row in [n + '1',n + '2',n + '3',n + '4',n + '5',n + '6',n + '7',n + '8']:
                    p = plate.cell(col + row).value
                    if p != None:
                        extra[p] = extra.get(p, 0) + 1
                    else:
                        break
            N += 1
        return extra
    
def read_qPCR(data):
    data_cols = {}
    for c in range(len(data.column_dimensions) + 2):
        data_name = data.cell(row = 0, column = c).value
        data_cols[data_name] = c
    if 'Well' not in data_cols.keys() or all(['Threshold Cycle ( C(t) )' not in data_cols.keys(), 'Cq' not in data_cols.keys()]):
        print('"Well" or "Threshold Cycle ( C(t) )/Cq" field not found in "data" sheet!\nRedo the sheet again following the following rules:\n0. Direct copy & paste all "Quantification Data"\n (or select at least "Well" and "Threshold Cycle ( C(t) )" columns into the sheet.\n1. Do not leave empty rows on top of sheet.\n2. Do not delete the field name row.\n3. Do not leave more than two empty columns on the left.n/4. Do not edit, delete, or rearrange anything after pasting.')
        print('*' * 20)
        print()
        return False
    Well_Ct = []
    if 'Cq' not in data_cols.keys():
        for r in range(1, len(data.row_dimensions)):
            Well_Ct.append([data.cell(row = r, column = data_cols['Well']).value, data.cell(row = r, column = data_cols['Threshold Cycle ( C(t) )']).value])
    else:
        for r in range(1, len(data.row_dimensions)):
            Well_Ct.append([data.cell(row = r, column = data_cols['Well']).value, data.cell(row = r, column = data_cols['Cq']).value])
    return Well_Ct
        
def if_one(Group):
    big_gap = 0
    last_one = 0
    N = len(Group)
    if N == 1:
        i = Group[0]
        if i[-1] < 30:
            i.append(1)
        else:
            i.append(0)
        return
    for i in range(N - 1):
        a = Group[i + 1][-1]
        b = Group[i][-1]
        dif = a - b
        if dif > big_gap:
            big_gap = dif
            last_one = i
    if i == 0 or N - i == 2: #To tell Single or all 1/0 case
        if big_gap < 2: #All 1/0
            if Group[0][-1] > 32:
                last_one = -1
            if Group[-1][-1] <= 32:
                last_one = N
    for i in range(N):
        if i <= last_one:
            Group[i].append(1)
        else:
            Group[i].append(0)

def sort(x):
    return (str(x[0]), x[-1], str(x[1]))

def Match_qPCR(data, plate_dict):
    old_Cts = None
    Cts = read_qPCR(data) #List of [Well, Ct]
    if Cts == False:
        return False
    a = input('How many cycles did you run? (default = 40)')
    try:
        Max_Ct = int(a) + 1
    except:
        Max_Ct = 41
    if 'plate_match' in wb.get_sheet_names():
        ask = input('"plate_match" sheet exists, do you want to append new results or overwrite the sheet? (a/o, default = a)')
        if ask == 'o' or ask == 'O':
            temp = wb.get_sheet_by_name('plate_match')
            wb.remove_sheet(temp)
        else:
            ws = wb.get_sheet_by_name('plate_match')
            old_Cts = []
            for r in range(1, len(ws.row_dimensions)):
                Ct = []
                Ct.append(ws.cell(row = r, column = 0).value)
                Ct.append(ws.cell(row = r, column = 1).value)
                Ct.append('o_' + ws.cell(row = r, column = 2).value)
                Ct.append(ws.cell(row = r, column = 3).value)
                Ct.append(ws.cell(row = r, column = 4).value)
                old_Cts.append(Ct) #old_Cts is a list of [primer, sample, well, Ct, IF]
            wb.remove_sheet(ws)
    for Ct in Cts:
        if Ct[1] == 'N/A':
            Ct[1] = Max_Ct
        Ct.insert(0,plate_dict[Ct[0]][1])
        Ct.insert(0,plate_dict[Ct[1]][0]) #Cts becomes list of [First primer pair, Tube no., Well, Ct]
    Cts_copy = Cts[:]
    for Ct in Cts_copy:
        if Ct[0] == None or Ct[1] == None:
            Cts.remove(Ct)
    Cts.sort(key = sort)
    Group = []
    p = Cts[0][0]
    for Ct in Cts:
        if Ct[0] == p:
            Group.append(Ct)
        else:
            if_one(Group)
            Group = [Ct]
            p = Ct[0]
    if_one(Group)
    match_qPCR = wb.create_sheet(0, 'plate_match')
    match_qPCR.cell(row = 0, column = 0).value = 'Target'
    match_qPCR.cell(row = 0, column = 1).value = 'Sample'
    match_qPCR.cell(row = 0, column = 2).value = 'Well'
    match_qPCR.cell(row = 0, column = 3).value = 'Ct'
    match_qPCR.cell(row = 0, column = 4).value = 'IF'
    r = 1
    if old_Cts:
        for i in range(len(old_Cts)):
            match_qPCR.cell(row = r, column = 0).value = old_Cts[i][0]
            match_qPCR.cell(row = r, column = 1).value = old_Cts[i][1]
            match_qPCR.cell(row = r, column = 2).value = old_Cts[i][2]
            match_qPCR.cell(row = r, column = 3).value = old_Cts[i][3]
            match_qPCR.cell(row = r, column = 4).value = old_Cts[i][4]
            r += 1
    for i in range(len(Cts)):
        match_qPCR.cell(row = r, column = 0).value = Cts[i][0]
        match_qPCR.cell(row = r, column = 1).value = Cts[i][1]
        match_qPCR.cell(row = r, column = 2).value = Cts[i][2]
        match_qPCR.cell(row = r, column = 3).value = Cts[i][3]
        match_qPCR.cell(row = r, column = 4).value = Cts[i][4]
        r += 1
    try:
        wb.save(F1)
        print('Now open "' + F1 + '" to review the matched results.')
        print('Only edit "if" field, do not delete columns.')
        print('Only delete rows when deemed invalid.')
        print('Do not leave blank rows inbetween ones with values.')
        return True
    except PermissionError:
        print('Please close "' + F1 + '" for file writing!')
        print('*' * 20)
        input()
        print()

def calculate_reagents(plate):
    plate_read = read_plate(plate)
    needed = {}
    for well in plate_read:
        p = plate_read[well][0]
        s = plate_read[well][1]
        if p != None and s != None:
            needed[p] = needed.get(p, 0) + 1
    needed_list = []
    for i in needed:
        needed_list.append((str(i), (needed[i] + needed[i] // 16) / 10))# the addition of 1/16 is adjustment for pippeting error
    needed_list.sort()
    Extra = read_plate(plate, extra = True)
    Extra_need = []
    for p in Extra:
        if p in needed:
            needed[p] += Extra[p]
    for i in Extra:
        if i in needed:
            Extra_need.append((str(i) + '(ex)', (needed[i] + needed[i] // 16) / 10))
    Extra_need.sort()
    needed_list += Extra_need
    needed_list.sort(key = lambda x: x[1], reverse = True)
    c = 0
    plate.cell(row = 10, column = c).value = 'Unit'
    plate.cell(row = 11, column = c).value = '2X mix'
    plate.cell(row = 12, column = c).value = 'P/P'
    plate.cell(row = 13, column = c).value = 'H2O'
    plate.cell(row = 14, column = c).value = 'Target'
    curr_v = 0
    for i in needed_list:
        if i[1] != curr_v:
            c += 1
            plate.cell(row = 10, column = c).value = str(i[1]) + 'X'
            plate.cell(row = 11, column = c).value = i[1] * 25
            plate.cell(row = 12, column = c).value = i[1] * 3
            plate.cell(row = 13, column = c).value = i[1] * 12
            plate.cell(row = 14, column = c).value = str(i[0])
            curr_v = i[1]
        else:
            plate.cell(row = 14, column = c).value = (str(plate.cell(row = 14, column = c).value) + '; ' + str(i[0]))
    try:
        wb.save(F1)
        print('Now open "' + F1 + '" to view the recipe table.')
        print()
    except PermissionError:
        print('Please close "' + F1 + '" for file writing!')
        print('*' * 20)
        print()
        return False
        
def read_GT_sheet():
    try:
        GTB = load_workbook(F2, use_iterators = True)
        print('"' + F2 + '" is used.')
        file_name = input('Enter file path and name if different.\n(To continue without change, just press enter.)\n>')
        if file_name:
            try:
                GTB = load_workbook(file_name, use_iterators = True)
            except:
                print(file_name + 'is missing!')
                print()
                return False
    except:
        print('"' + F2 + '" is missing!')
        print()
        file_name = input('Enter file path and name if different:')
        print()
        try:
            GTB = load_workbook(file_name, use_iterators = True)
        except:
            print(file_name + 'is missing!')
            print()
            return False
    TS_name = GTB.get_sheet_names()[1]
    if input('Is "' + TS_name + '" the sheet to work on ? (y/n)') not in ['Y', 'y', '']:
        TS_name = input('What is the sheet name?')
        if TS_name not in GTB.get_sheet_names():
            print('The sheet "' + TS_name + '" is not in the file!')
            print()
            return False
    TS1 = GTB.get_sheet_by_name(TS_name)
    return TS1

def read_match_qPCR(ws):
    Dict = {}
    for r in range(1, len(ws.row_dimensions)):
        target = ws.cell(row = r, column = 0).value
        sample = ws.cell(row = r, column = 1).value
        IF = ws.cell(row = r, column = 4).value
        Dict[(sample, target)] = IF
    return Dict

def sort_int_str(list):
    Str = []
    Non = []
    for i in list:
        if type(i) == str:
            Str.append(i)
        else:
            Non.append(i)
    Str.sort()
    Non.sort()
    return Non + Str

def match_Toes(match_qPCR):
    targets = []
    if 'Toe_match' not in wb.get_sheet_names():
        TS1 = read_GT_sheet()
        if TS1 == False:
            return False
        match_toes = wb.create_sheet(0, 'Toe_match')
        Samples = read_cols(TS1, two = False)
        STP_dict_existing = None
    else:
        match_toes = wb.get_sheet_by_name('Toe_match')
        Samples = []
        for r in range(1, len(match_toes.row_dimensions)):
            Samples.append(match_toes.cell(row = r, column = 0).value)
        for c in range(1, len(match_toes.column_dimensions)):
            targets.append(match_toes.cell(row = 0, column = c).value)
        STP_dict_existing = dict()
        for r in range(1, len(match_toes.row_dimensions)):
            for c in range(1, len(match_toes.column_dimensions)):
                val = match_toes.cell(row = r, column = c).value
                if val != None:
                    STP_dict_existing[(match_toes.cell(row = r, column = 0).value, match_toes.cell(row = 0, column = c).value)] = val
        wb.remove_sheet(match_toes)
        match_toes = wb.create_sheet(0, 'Toe_match')
    STP_dict = read_match_qPCR(match_qPCR) #sample, target pair dict as (sample, target): 1/0
    if STP_dict_existing:
        STP_dict_new = dict(list(STP_dict_existing.items()) + list(STP_dict.items()))
        STP_dict = STP_dict_new
    add_Samples = set()
    add_targets = set()
    for (s, t) in STP_dict:
        if s not in Samples:
            add_Samples.add(s)
        if t not in targets:
            add_targets.add(t)
    add_Samples = list(add_Samples)
    add_targets = list(add_targets)
    add_Samples1 = sort_int_str(add_Samples)
    add_targets1 = sort_int_str(add_targets)
    Samples.extend(add_Samples1)
    targets.extend(add_targets1)
    match_toes.cell(row = 0, column = 0).value = 'Tube#'
    S_axis = {}
    t_axis = {}
    for i in range(len(Samples)):
        r = i + 1
        s = Samples[i]
        match_toes.cell(row = r, column = 0).value = s
        S_axis[s] = r
    for i in range(len(targets)):
        c = i + 1
        t = targets[i]
        match_toes.cell(row = 0, column = c).value = t
        t_axis[t] = c
    for (s, t) in STP_dict:
        match_toes.cell(row = S_axis[s], column = t_axis[t]).value = STP_dict[(s, t)]
    try:
        wb.save(F1)
        print('Now open "' + F1 + '" to view the sample matched results.\n (In "Toe_match" sheet)')
        print()
        return True
    except PermissionError:
        print('Please close "' + F1 + '" for file writing!')
        print('*' * 20)
        print()
        return False

def start():
    global wb, litters, F1
    done = False
    print('Welcome to Toe Sheet Generator!')
    while done == False:
        Job = job()
        if Job == 'c' or Job == 'C':
            done = True
            continue
        try:
            wb = load_workbook(F1)
        except:
            print('"' + F1 + '" does not exist!')
            F1 = input('Enter file path and name if different:')
            try:
                wb = load_workbook(F1)
            except:
                print('"' + F1 + '" does not exist!')
        if Job == 1:
            if 'Litters for GT' in wb.get_sheet_names():
                print('"Litters for GT" sheet already exist, to redo it, delete it first, or proceed with the existing sheet and choose option 2.')
                print('*' * 20)
                print()
                continue
            collect_litters()# collect info of all litters in current cross
            mixed_index(litters) # findout which litters are in the same cage
            filter_litters(litters) # keep only ungenotyped and old enough litters
            special_case(litters) # find additional (VISUAL) alleles
            litters_sort(litters) # sort list based on criterias
            GT1 = wb.create_sheet(index = 0, title = 'Litters for GT')
            make_list(GT1)
            try:
                wb.save(F1)
                print('Now open "' + F1 + '" to review the list.')
                print('*Make necessary changes, then save and close file to continue.')
                print('*Do not delete field name row, or nothing will be changed.')
                print('*Do not edit cells in "litterID" column.')
                print('*Columns can be deleted or swapped.')
                print('*If "litterID" column is deleted, no change will be made.')
                print('*Edits of "realsize", "first", "mixed_index" cells will be updated.')
                print('*Change of order or deletion of litters will be updated.')
                print('*Delete "Litters for GT" sheet if intend to redo the list again.')
                print('*No blank columns or blank rows between ones with values, or error will occur!')
                print()
                print()
            except PermissionError:
                print('Please close' + F1 + 'for file writing!')
                litters = []
                print()
        elif Job == 2:
            GT1 = wb.get_sheet_by_name('Litters for GT')
            if GT1 == None:
                print('"Litters for GT" sheet has not been generated!\nGo through option 1 first!')
                print('*' * 20)
                print()
            else:
                if litters == []:
                    collect_litters()# collect info of all litters in current cross
                    mixed_index(litters =  litters) # findout which litters are in the same cage
                    special_case(litters = litters) # find additional (VISUAL) alleles
                if list_updater(GT1) == False:
                    print('Internal empty column or row in "Litter for GT" encountered, revise "Litters for GT" sheet;\n or delete it and rerun option 1.')
                    print('*' * 20)
                    print()
                else:
                    generate_TS()
                    
        elif Job == 3:
            TS1 = read_GT_sheet()
            if TS1 != False:
                generate_plate(TS1)
        elif Job == 4:
            print('Will read "plate" sheet, if use different sheet, enter name:')
            print('(Just press "Enter" if no change)')
            plate_name = input('>')
            if plate_name == '':
                plate_name = 'plate'
            plate = wb.get_sheet_by_name(plate_name)
            if plate == None:
                print('"' + plate_name + '" sheet is missing!')
                print('*' * 20)
                print()
            else:
                calculate_reagents(plate)
            
        elif Job == 5:
            print('Will read "plate" sheet for extracting primer/sample/well information,\n if use different sheet, enter name:')
            print('(Just press "Enter" if no change)')
            plate_name = input('>')
            if plate_name == '':
                plate_name = 'plate'
            print('Will read "data" sheet for extracting qPCR results,\n if use different sheet, enter name:')
            print('(Just press "Enter" if no change)')
            data_name = input('>')
            if data_name == '':
                data_name = 'data'
            plate = wb.get_sheet_by_name(plate_name)
            data = wb.get_sheet_by_name(data_name)
            if data == None:
                if data_name == 'data':
                    print('"data" sheet has not been generated!\nCopy & paste qPCR results into "data" sheet!')
                else:
                    print('"' + data_name + '" sheet is missing!')
                print('*' * 20)
                print()
            elif plate == None:
                print('"' + plate_name + '" sheet is missing!')
                print('*' * 20)
                print()
            else:
                Match_qPCR(data, read_plate(plate))
        elif Job == 6:
            print('Will read "plate_match" sheet, if use different sheet, enter name:')
            print('(Just press "Enter" if no change)')
            match_name = input('>')
            if match_name == '':
                match_name = 'plate_match'
            match_qPCR = wb.get_sheet_by_name(match_name)
            if match_qPCR == None:
                print('"' + match_name +'" sheet is missing!')
                print('Run Option 5 to generate it first!')
                print('*' * 20)
                print()
                continue
            match_Toes(match_qPCR)
        input('Press "Enter" to continue...')
    print('-' * 20 + '\nBye!')

start()







